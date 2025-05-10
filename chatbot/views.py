from django.shortcuts import render, redirect
from django.http import JsonResponse,FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.forms import modelformset_factory
from .chatbot_logic import get_ai_response
from .models import KnowledgeBase, KnowledgeBaseFile
from .forms import CreateKBForm, KnowledgeBaseFileForm
from django.contrib import messages
from .report_generator1 import  generate_chat_response, create_vector_store1,generate_excel_report_from_kb,generate_qa_report_from_kb,get_retriever_for_kb
import os
import logging
from django.shortcuts import get_object_or_404,redirect
import json
from openpyxl import load_workbook
from datetime import datetime

logger = logging.getLogger(__name__)


@csrf_exempt
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Load users from JSON file
        json_file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'users.json')
        
        try:
            with open(json_file_path, 'r') as f:
                users_data = json.load(f)
            
            # Check credentials
            authenticated = False
            user_info = None
            
            for user in users_data.get('users', []):
                if user['username'] == username and user['password'] == password:
                    authenticated = True
                    user_info = user
                    break
            
            if authenticated:
                # Store user info in session
                request.session['user'] = {
                    'username': user_info['username'],
                    'full_name': user_info['full_name'],
                    'role': user_info['role'],
                    'is_authenticated': True
                }
                messages.success(request, f"Welcome back, {user_info['full_name']}!")
                return redirect('index')
            else:
                messages.error(request, "Invalid username or password")
                
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
    
    return render(request, 'login.html')

def logout_view(request):
    # Clear the session
    if 'user' in request.session:
        del request.session['user']
    messages.success(request, "You have been logged out successfully")
    return redirect('login')


def chat(request):
    if request.method == 'POST':
        user_input = request.POST.get('user_input')
        company_name = request.POST.get('company', '')
        kb_id = request.POST.get('kb_id')
        
        # Add proper validation and logging
        if not user_input:
            return JsonResponse({'response': "Please enter a question."})
            
        if not company_name:
            return JsonResponse({'response': "Please select a company before asking questions."})
            
        if not kb_id:
            return JsonResponse({'response': "Please select a knowledge base before asking questions."})
            
        # Look up the knowledge base name from the ID
        try:
            from chatbot.models import KnowledgeBase
            kb_name = kb_id
            
            # Now call generate_chat_response with the actual kb_name
            ai_response = generate_chat_response(user_input, company_name, kb_name)
            return JsonResponse({'response': ai_response})
        except KnowledgeBase.DoesNotExist:
            return JsonResponse({'response': f"Knowledge base with ID {kb_id} not found. Please select a valid knowledge base."})
        except Exception as e:
            logger.error(f"Error in chat view: {e}")
            return JsonResponse({'response': "An error occurred while processing your request. Please try again."})
    else:
        return render(request, 'chatbot.html')

def index(request):
    """
    View function for the home page of the chatbot.
    """
    # Get list of companies from the database
    from .models import Company
    companies = list(Company.objects.values_list('name', flat=True))
    
    # Default if no companies exist
    if not companies:
        # Add default company if none exist
        default_company = "Central Bank of Bahrain"
        Company.objects.create(name=default_company)
        companies = [default_company]

    context = {
        'companies': companies,
    }
    return render(request, 'index.html', context)

def add_company(request):
    """Add a new company to the system"""
    if request.method == 'POST':
        try:
            # Get company name from form
            company_name = request.POST.get('company_name', '').strip()
            
            if not company_name:
                return JsonResponse({'success': False, 'message': 'Company name is required'})
            
            # Check if company already exists
            from .models import Company
            if Company.objects.filter(name=company_name).exists():
                return JsonResponse({'success': False, 'message': 'Company already exists'})
            
            # Create new company (directories are created in the save method)
            company = Company.objects.create(name=company_name)
            
            return JsonResponse({
                'success': True, 
                'message': f'Company "{company_name}" added successfully',
                'company_name': company_name
            })
            
        except Exception as e:
            logger.error(f"Error adding company: {e}")
            return JsonResponse({'success': False, 'message': f"Error: {str(e)}"})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


def create_knowledge_base(request):
    """
    View function to create a new knowledge base
    """
    selected_company = request.GET.get('company', 'Default Company')

     # Ensure company exists
    from .models import Company
    if not Company.objects.filter(name=selected_company).exists():
        Company.objects.create(name=selected_company)
    
    if request.method == 'POST':
        kb_form = CreateKBForm(request.POST, company=selected_company)
        
        if kb_form.is_valid():
            try:
                # Create the knowledge base
                kb = kb_form.save()


                 # Add the company to the context for the response
                new_company = selected_company
                
                # Handle multiple file uploads
                files = request.FILES.getlist('files')
                for f in files:
                    KnowledgeBaseFile.objects.create(knowledge_base=kb, file=f)
                
                kb_name = kb.name
                
                # Sanitize the kb_name for use as collection name
                import re
                sanitized_kb_name = re.sub(r'[^a-zA-Z0-9_-]', '_', kb_name)
                # Ensure it starts and ends with alphanumeric character
                if not sanitized_kb_name[0].isalnum():
                    sanitized_kb_name = 'kb_' + sanitized_kb_name
                if len(sanitized_kb_name) > 0 and not sanitized_kb_name[-1].isalnum():
                    sanitized_kb_name = sanitized_kb_name + '1'
                # Ensure minimum length of 3
                if len(sanitized_kb_name) < 3:
                    sanitized_kb_name = sanitized_kb_name + '_db'
                # Ensure maximum length of 63
                sanitized_kb_name = sanitized_kb_name[:63]
                
                # Create vector store
                try:
                    from .report_generator1 import create_vector_store1
                    retriever = create_vector_store1(files, sanitized_kb_name, selected_company)
                    if retriever is None:
                        messages.error(request, "Failed to create vector store. Check logs for details.")
                    else:
                        messages.success(request, f"Knowledge base '{kb.name}' created successfully with {len(files)} files for company '{new_company}'.")
                except Exception as e:
                    logger.error(f"Error creating vector store: {e}")
                    messages.error(request, f"Error creating knowledge base: {str(e)}")


                
                return redirect('index')
            except Exception as e:
                messages.error(request, f"Error creating knowledge base: {str(e)}")
    else:
        kb_form = CreateKBForm(company=selected_company)
    
    context = {
        'form': kb_form,
        'selected_company': selected_company
    }
    return render(request, 'create_kb.html', context)


def get_knowledge_bases(request):
    """AJAX view to get knowledge bases for a company"""
    company = request.GET.get('company', '')
    
    print(f"get_knowledge_bases called with company: '{company}'")
    
    if company:
        folder_path = "media/vector_stores/" + company
        
        # Check if directory exists
        if not os.path.exists(folder_path):
            # Create directory if it doesn't exist
            os.makedirs(folder_path)
            return JsonResponse({'knowledge_bases': []})
            
        file_names = os.listdir(folder_path)
        result = {'knowledge_bases': file_names}
        print(f"Returning {len(file_names)} knowledge bases: {result}")
        return JsonResponse(result)
    else:
        print("No company provided, returning empty list")
        return JsonResponse({'knowledge_bases': []})



def generate_report(request):
    """
    View function to generate a report for a knowledge base
    """
    if request.method == 'POST':
        try:
            # Parse JSON data
            data = json.loads(request.body)
            company_name = data.get('company')
            kb_name = data.get('kb_name')
            report_type = data.get('report_type', 'excel')  
            prompt_type = data.get('prompt_type')  # Default to 'prompt_1' if not provided

            # Validate input
            if not company_name or not kb_name:
                return JsonResponse({
                    'success': False,
                    'message': 'Company name and knowledge base name are required'
                })
            
            # Initialize result
            result = {
                'success': False,
                'message': 'Report generation failed'
            }
            
            # Generate the requested report type
            if report_type == 'excel':
                logger.info(f"Generating Excel report for {company_name}/{kb_name} using {prompt_type}")
                ui_prompt_selection = data.get('prompt_type') 
                backend_prompt_type = "prompt_2" if ui_prompt_selection == "prompt_2" else "prompt_1"
                report_path = generate_excel_report_from_kb(company_name, kb_name, prompt_type=backend_prompt_type)
                logger.info(f"Report path: {report_path}")
                
                if report_path:
                    result = {
                        'success': True,
                        'message': 'Excel report generated successfully',
                        'report_path': report_path
                    }

                    # After successfully generating the report
                    if 'generated_reports' not in request.session:
                        request.session['generated_reports'] = []

                     # Add the new report to the session
                    request.session['generated_reports'].append({
                        'company': company_name,
                        'kb_name': kb_name,
                        'report_type': report_type,
                        'prompt_type': prompt_type,
                        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'report_path': report_path
                    })

                     # Save session changes
                    request.session.modified = True


                     # Still save to database for permanent storage
                    from .models import Report
                    Report.objects.create(
                        company=company_name,
                        kb_name=kb_name,
                        report_path=report_path,
                        report_type=report_type,
                        prompt_type=prompt_type
                    )
                else:
                    logger.error(f"Failed to generate Excel report for {company_name}/{kb_name}")
                    result = {
                            'success': False,
                            'message': 'Failed to generate Excel report'
                    }
            
            
           
            
            # Return the result
            return JsonResponse(result)
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            return JsonResponse({
                'success': False,
                'message': f"Error: {str(e)}"
            })
    else:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request method'
        }, status=405)    

# Add this new function

def get_reports(request):
    """
    View function to get all generated reports
    """
    try:
        from .models import Report
        reports = Report.objects.all()
        
        reports_data = []
        for report in reports:
            reports_data.append({
                'id': report.id,
                'company': report.company,
                'kb_name': report.kb_name,
                'report_type': report.report_type,
                'prompt_type': getattr(report, 'prompt_type'),  # Default to 'prompt_1' if not set
                'created_at': report.created_at.strftime('%Y-%m-%d %H:%M'),
                'report_path': report.report_path
            })
        
        return JsonResponse({'success': True, 'reports': reports_data})
    except Exception as e:
        logger.error(f"Error fetching reports: {e}")
        return JsonResponse({
            'success': False,
            'message': f"Error: {str(e)}"
        })
    
def get_report_content(request, report_path):
    """
    View function to read Excel report content and return it as JSON
    """
    try:
        # Check if file exists
        if not os.path.exists(report_path):
            return JsonResponse({
                'success': False,
                'message': 'Report file not found'
            }, status=404)
        
        # Check if file is an Excel file
        if not report_path.endswith('.xlsx'):
            return JsonResponse({
                'success': False,
                'message': 'File is not an Excel report'
            }, status=400)
        
        # Load the workbook
        wb = load_workbook(filename=report_path)
        ws = wb.active
        
        # Extract the company name from the title cell (A1)
        title = ws['A1'].value
        company_name = title.split(' - ')[0] if ' - ' in title else 'Company'

        header_d3 = ws['D3'].value or ""
        header_e3 = ws['E3'].value or ""

        # Your existing detection logic:
        if "AI" in header_d3 and "Retrieved" in header_e3:
            report_type = "qa_excel"
        elif "AI" in header_d3 and "Faithfulness" in header_e3:
            report_type = "qa_excel_evaluated"
        else:
            report_type = "excel"
            
        if report_type == "qa_excel_evaluated":
            # Process as a Q&A report
            report_data = {
                'title': title,
                'qa_data': {}
            }
             # Start from row 4 (after headers)
            current_category = None
            for row in range(4, ws.max_row + 1):
                # Check if this is a category row (merged cells)
                merged_cell_ranges = [str(cell_range) for cell_range in ws.merged_cells.ranges]
                current_cell = f'A{row}:H{row}'
                
                if current_cell in merged_cell_ranges:
                    # Start a new category
                    current_category = ws[f'A{row}'].value
                    report_data['qa_data'][current_category] = []
                else:
                    # Regular question row
                    question_type = ws[f'A{row}'].value
                    question = ws[f'B{row}'].value
                    analyst_answer = ws[f'C{row}'].value
                    ai_answer = ws[f'D{row}'].value
                    faithfulness = ws[f'E{row}'].value
                    relevancy = ws[f'F{row}'].value
                    context_precision = ws[f'G{row}'].value
                    answer_correctness = ws[f'H{row}'].value
                    
                    if question and (analyst_answer or ai_answer):
                        report_data['qa_data'][current_category].append({
                            'question': question,
                            'analyst_answer': analyst_answer or "Not provided",
                            'ai_answer': ai_answer or "Not provided",
                            'faithfulness': faithfulness or "N/A",
                            'relevancy': relevancy or "N/A",
                            'context_precision': context_precision or "N/A",
                            'answer_correctness': answer_correctness or "N/A"
                        })
            
            return JsonResponse({
                'success': True,
                'report_data': report_data,
                'report_type': report_type
            })
        
        elif report_type == "qa_excel":
            # Process as a QA Excel report
            report_data = {
                'title': title,
                'qa_data': {}
            }
            
            # Start from row 4 (after headers)
            current_category = None
            for row in range(4, ws.max_row + 1):
                # Check if this is a category row (merged cells)
                merged_cell_ranges = [str(cell_range) for cell_range in ws.merged_cells.ranges]
                current_cell = f'A{row}:H{row}'
                
                if current_cell in merged_cell_ranges:
                    # Start a new category
                    current_category = ws[f'A{row}'].value
                    report_data['qa_data'][current_category] = []
                else:
                    # Regular question row
                    question_type = ws[f'A{row}'].value
                    question = ws[f'B{row}'].value
                    analyst_answer = ws[f'C{row}'].value
                    ai_answer = ws[f'D{row}'].value
                    retrieved_answer = ws[f'E{row}'].value
                    
                    if question and (analyst_answer or ai_answer):
                        report_data['qa_data'][current_category].append({
                            'question': question,
                            'analyst_answer': analyst_answer or "Not provided",
                            'ai_answer': ai_answer or "Not provided",
                            'retrieved_answer': retrieved_answer or "N/A"
                        })
            
            return JsonResponse({
                'success': True,
                'report_data': report_data,
                'report_type': report_type
            })

        else:
            # Initialize data structure
            report_data = {
                'title': title,
                'company': company_name,
                'categories': []
            }
            
            # Process the rows and collect the data
            current_category = None
            category_questions = []
            
            # Start from row 4 (after headers)
            for row in range(4, ws.max_row + 1):
                # Check if this is a category row (merged cells)
                merged_cell_ranges = [str(cell_range) for cell_range in ws.merged_cells.ranges]
                current_cell = f'A{row}:C{row}'
                
                if current_cell in merged_cell_ranges:
                    # If we already have a category, add it to the report data
                    if current_category and category_questions:
                        report_data['categories'].append({
                            'name': current_category,
                            'questions': category_questions
                        })
                    
                    # Start a new category
                    current_category = ws[f'A{row}'].value
                    category_questions = []
                else:
                    # Regular question row
                    question = ws[f'B{row}'].value
                    answer = ws[f'C{row}'].value
                    
                    if question and answer:
                        category_questions.append({
                            'question': question,
                            'answer': answer
                        })
            
            # Add the last category if there is one
            if current_category and category_questions:
                report_data['categories'].append({
                    'name': current_category,
                    'questions': category_questions
                })
            
            return JsonResponse({
                'success': True,
                'report_data': report_data,
                'report_type': report_type
            })
            
    except Exception as e:
            logger.error(f"Error reading report content: {e}")
            return JsonResponse({
                'success': False,
                'message': f"Error reading report content: {str(e)}"
            }, status=500)


    

def download_report(request, report_path):
    """
    View function to download a generated report
    """
    try:
        # Validate the file exists
        if not os.path.exists(report_path):
            return JsonResponse({
                'success': False,
                'message': 'Report file not found'
            }, status=404)
        
        # Get the file extension
        file_extension = os.path.splitext(report_path)[1].lower()
        
        # Set the content type based on file extension
        content_types = {
            '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
            '.pdf': 'application/pdf',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        content_type = content_types.get(file_extension, 'application/octet-stream')

         # Get view parameter - if true, try to display in browser
        view = request.GET.get('view', 'false').lower() == 'true'
        
        # Create a FileResponse
        response = FileResponse(open(report_path, 'rb'), content_type=content_type)

         # Set the content disposition based on the view parameter
        if view and file_extension == '.pdf':
            # PDF files can be displayed in-browser
            response['Content-Disposition'] = f'inline; filename="{os.path.basename(report_path)}"'
        else:
            # All other files should be downloaded
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(report_path)}"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error downloading report: {e}")
        return JsonResponse({
            'success': False,
            'message': f"Error: {str(e)}"
        }, status=500)
    

def generate_qa_report(request):
    """
    View function to generate a question-based report for a knowledge base
    """
    if request.method == 'POST':
        try:
            # Parse JSON data
            data = json.loads(request.body)
            company_name = data.get('company')
            kb_name = data.get('kb_name')
            ground_truths_path = data.get('ground_truths_path')
            prompt_type = data.get('prompt_type')
            
            logger.info(f"Generating QA report for {company_name}/{kb_name} using {prompt_type}")
            logger.info(f"Ground truths path: {ground_truths_path}")
            
            # Validate input
            if not company_name or not kb_name:
                logger.error("Missing company name or KB name")
                return JsonResponse({
                    'success': False,
                    'message': 'Company name and knowledge base name are required'
                })
            
            # Check if ground truths file exists if provided
            if ground_truths_path:
                if not os.path.exists(ground_truths_path):
                    logger.warning(f"Ground truths file not found at {ground_truths_path}")
                    ground_truths_path = None
                else:
                    logger.info(f"Ground truths file confirmed at {ground_truths_path}")
            
            # Get the retriever for this knowledge base
            retriever = get_retriever_for_kb(company_name, kb_name)
            if not retriever:
                logger.error(f"Failed to get retriever for {company_name}/{kb_name}")
                return JsonResponse({
                    'success': False,
                    'message': f'Knowledge base not found for {company_name}/{kb_name}'
                })
            
            
            # Generate the report using the predefined question set
            logger.info("Retrieved retriever, now generating report...")
            ui_prompt_selection = data.get('prompt_type')
            backend_prompt_type = "prompt_2" if ui_prompt_selection == "prompt_2" else "prompt_1"
            report_path = generate_qa_report_from_kb(company_name, kb_name, retriever,ground_truths_path, prompt_type=backend_prompt_type)
            
            if not report_path:
                logger.error("Failed to generate report, return path was None")
                return JsonResponse({
                    'success': False,
                    'message': 'Failed to generate QA report'
                })
                
            # Success - save to database and return
            if report_path:
                logger.info(f"Report successfully generated at: {report_path}")
                report_db_type = 'qa_excel_evaluated' if ground_truths_path else 'qa_excel'
                from .models import Report
                Report.objects.create(
                    company=company_name,
                    kb_name=kb_name,
                    report_path=report_path,
                    report_type=report_db_type,
                    prompt_type=prompt_type
                )

                # Update Session for qa report
                if 'generated_reports' not in request.session:
                    request.session['generated_reports'] = []
                request.session['generated_reports'].append({
                    'company': company_name,
                    'kb_name': kb_name,
                    'report_type': report_db_type, # Use the same report_type
                    'prompt_type': prompt_type,
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'report_path': report_path
                })
                request.session.modified = True
            
            return JsonResponse({
                'success': True,
                'message': 'Q&A Report generated successfully',
                'report_path': report_path
            })
            
        except Exception as e:
            logger.error(f"Error generating Q&A report: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'message': f"Error: {str(e)}"
            })
    else:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request method'
        }, status=405)
    
def get_session_reports(request):
    """Return only reports generated in the current session."""
    try:
        # Get reports only from the current session
        session_reports = request.session.get('generated_reports', [])
        
        # Format reports for the frontend
        reports = []
        for report in session_reports:
            reports.append({
                'company': report.get('company', ''),
                'kb_name': report.get('kb_name', ''),
                'report_type': report.get('report_type', ''),
                'created_at': report.get('created_at', ''),
                'report_path': report.get('report_path', ''),
                'prompt_type': report.get('prompt_type')  # Default to 'prompt_1' if not set,
            })
            
        return JsonResponse({
            'success': True,
            'reports': reports
        })
    except Exception as e:
        logger.error(f"Error getting session reports: {e}")
        return JsonResponse({
            'success': False,
            'message': 'Error retrieving reports'
        })
    

def upload_ground_truths(request):
    """
    View function to upload ground truths file
    """
    if request.method == 'POST':
        try:
            # Get form data
            file_obj = request.FILES.get('file')
            company_name = request.POST.get('company')
            kb_name = request.POST.get('kb_name')
            
            if not file_obj or not company_name or not kb_name:
                return JsonResponse({
                    'success': False,
                    'message': 'Missing file, company name, or KB name'
                })
            
            # Ensure the file is a markdown or text file
            file_extension = os.path.splitext(file_obj.name.lower())[1]
            if file_extension not in ['.md', '.txt']:
                return JsonResponse({
                    'success': False,
                    'message': 'Only markdown (.md) or text (.txt) files are supported for ground truths'
                })
            
            # Get directory for the ground truths
            from .report_generator1 import get_kb_dir, ensure_dir_exists
            kb_dir = get_kb_dir(company_name, kb_name)
            gt_dir = os.path.join(kb_dir, "ground_truths")
            ensure_dir_exists(gt_dir)
            
            # Save the file
            file_path = os.path.join(gt_dir, file_obj.name)
            with open(file_path, 'wb+') as destination:
                for chunk in file_obj.chunks():
                    destination.write(chunk)
            
            return JsonResponse({
                'success': True,
                'message': 'Ground truths file uploaded successfully',
                'file_path': file_path
            })
            
        except Exception as e:
            logger.error(f"Error uploading ground truths: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'message': f"Error: {str(e)}"
            })
    else:
        return JsonResponse({
            'success': False,
            'message': 'Invalid request method'
        }, status=405)


    
# Export the login and logout views explicitly
__all__ = [
    'login_view', 
    'logout_view', 
    'index', 
    'chat', 
    'create_knowledge_base', 
    'get_knowledge_bases',
    'add_company',
    'generate_report',
    'generate_qa_report',
    'download_report',
    'get_reports',
    'get_report_content',
    'upload_ground_truths',
    'get_session_reports'
]