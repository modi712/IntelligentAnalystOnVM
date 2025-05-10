import os
import uuid
import datetime
import logging
from pathlib import Path
from django.conf import settings
from uuid import uuid4
import shutil
from langchain_community.document_loaders import PyPDFLoader, CSVLoader
from langchain_huggingface.embeddings import HuggingFaceEmbeddings
from langchain.embeddings import SentenceTransformerEmbeddings
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import StrOutputParser, JsonOutputParser
from langchain_core.prompts import PromptTemplate
from langchain_chroma import Chroma
from langchain_core.pydantic_v1 import BaseModel, Field, validator
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.chains import LLMChain
import json
from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side 
import traceback
from datetime import datetime


# Configure logger
logger = logging.getLogger(__name__)

# Initialize constants
MEDIA_ROOT = settings.MEDIA_ROOT
REPORTS_DIR = os.path.join(MEDIA_ROOT, "reports")
VECTOR_STORES_DIR = os.path.join(MEDIA_ROOT, "vector_stores")

# Initialize embeddings and text splitter
embeddings = SentenceTransformerEmbeddings(model_name="all-MiniLM-L6-v2")
text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=100)

PROMPT_1 = """You are a professional business analyst creating a report.
Based on the following context information, provide a well-structured, professional response to the query.
Make your response comprehensive yet concise, suitable for a business presentation slide.

Important formatting requirements:
1. Start with a brief summary (1-2 sentences)
2. Use bullet points for key information (• format)
3. If providing data or statistics, present them clearly
4. For any multi-part answers, use numbered lists
5. Bold important terms or figures using markdown (**term**)
6. Use short paragraphs with clear spacing between points
7. Avoid long, dense paragraphs of text
{length_guidance}

Context: {context}
Query: {query}
Response:"""


PROMPT_2 = """You are a financial analyst with about 3 years of experience. Your job is to analyse companies and their busineses. All your responses should be aimed at a senior person who wants to get a well-structured, professional response to the query.
You are required to focus on the following key aspects:
- Develop an independent view
- Differentiate between fact and opinion. Triangulate based on all the information provided to you
- Focus a lot on actual facts rather than on just a story being told (by either the company or by analysts)
- Think of both the pros and cons and then come to a conclusion on what seems plausible
Make your response comprehensive yet concise, suitable for a business presentation slide.

Important formatting requirements:
1. Start with a brief summary (1-2 sentences)
2. Use bullet points for key information (• format)
3. If providing data or statistics, present them clearly
4. For any multi-part answers, use numbered lists
5. Bold important terms or figures using markdown (**term**)
6. Use short paragraphs with clear spacing between points
7. Avoid long, dense paragraphs of text
{length_guidance}

Context: {context}
Query: {query}
Response:"""

# Create directory structure functions
def ensure_dir_exists(path):
    """Ensure the directory exists, create if it doesn't."""
    if not os.path.exists(path):
        os.makedirs(path)
    return path


# Initialize LLM for report generation
try:
    from .api_key import GROQ_API_KEY  # Import API key
    from langchain_groq import ChatGroq as Groq
    # llm = Groq(api_key='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', model_name="llama-3.3-70b-versatile")
    llm = Groq(api_key='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', model_name="llama-3.3-70b-versatile")
except (ImportError, Exception) as e:
    logger.error(f"Failed to initialize LLM: {e}")
    llm = None

def get_company_dir(company_name):
    """Get company directory path and create if doesn't exist."""
    company_path = os.path.join(VECTOR_STORES_DIR, company_name)
    return ensure_dir_exists(company_path)

def get_kb_dir(company_name, kb_name):
    """Get knowledge base directory path and create if doesn't exist."""
    kb_path = os.path.join(get_company_dir(company_name), kb_name)
    return ensure_dir_exists(kb_path)




def generate_excel_report(company_name, kb_name, retriever, prompt_type="prompt_1"):
    """Generate an Excel report with answers to predefined questions"""
    logger.info(f"Generating Excel report for {company_name}/{kb_name}")
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Investment Analysis"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    category_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="CFCFCF", end_color="CFCFCF", fill_type="solid")
    category_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add report title
    ws.merge_cells('A1:C1')
    ws['A1'] = f"{company_name} - Investment Analysis Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add headers
    ws['A3'] = "Category"
    ws['B3'] = "Question"
    ws['C3'] = "Answer"
    
    for cell in ['A3', 'B3', 'C3']:
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 60
    
    # Define questions by category
    questions = {
        "Business Model": [
            "What is the TAM for each of the key revenue segments currently. Give me an INR number for each segment. take the above and compute total tam. Then compare with  current revenues for each segment and total.",
            "for the above, could you calculate the profit opportunity be each segment in INR currently and in the future",
            "what is the average capex for the company in INR. what are the average sales of the company in INR. calculate capex/sales"
        ],
        "Product and Technology": [
            "what is the key IP in the business - is it technology, logistics, operations or something else",
            "what is the average R&D expediture. What is the average revenues over the same period. Calculate the ratio of R&D to Revenues"
        ],
        "Customer/Value Proposition": [
            "Profile their typical customer. What value do they provide to the end customerr",
            "what is the competition to the company. How does the company differentiate vs them. Do they have a strong moat (rank 1-10, 10 being the strongest)"
        ],
        "Life Cycle/Competitive Analysis": [
            "where in the company lifecycle does the company stand",
            "which industry is the company operating. Is there lot of competition",
            "which industry is the company operating.. will the competition for the company increase or decrease in the coming 5 years"
        ],
        "Market View": [
            "are street analysts positive or negative on the company. What is their rationale",
            "who are the key shareholders. How much do promotors/founders own; how much is owned by institutional shareholders. Who are the key institutional shareholders"
        ],
        "Financials": [
            "is the company having fast revenue growth (greater than 20% per year over the past 2-3 years)",
            " Is the company making profits? If yes, are these profits growing. If not, is there a clear path to profitability"
        ],
        "Management Quality": [
            "who are the key maangement personnel. rate them for industry expertise",
            "who is the founder. Does he appear honest, trustworthy and ethical"
        ],
        "Regulatory Risk": [
            "Is there risk of governmental intervention"
        ],
        "Promotor Quality": [
            "who is the founder. Does he appear honest, trustworthy and ethical"
        ],
        "Key Risks": [
            "what are the key risks in the industry"
        ]
    }
    
    # Starting row for data
    row = 4
    
    # Iterate through categories and questions
    for category, category_questions in questions.items():
        # Add category row
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = category_font
        ws[f'A{row}'].fill = category_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left')
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = thin_border
        row += 1
        
        # Process each question in the category
        for question in category_questions:
            # Format the query for better results
            query = f"For {company_name}, {question}. Please provide a concise answer based on the available information."
            
            # Get answer from the retriever and LLM
            answer = run_query(query, retriever, prompt_type=prompt_type)
            
            # Add to worksheet
            ws[f'A{row}'] = ""  # No category in question rows
            ws[f'B{row}'] = question
            ws[f'C{row}'] = answer
            ws[f'C{row}'].alignment = Alignment(wrap_text=True, vertical='top')

            # Add these styles to make the text more readable
            ws[f'C{row}'].font = Font(name='Calibri', size=10)

            # Adjust row height for better readability
            ws.row_dimensions[row].height = max(18, min(40, 15 * (answer.count('\n') + 1)))
            
            # Apply borders
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = thin_border
                ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
    
    # Save the workbook
    excel_dir = os.path.join(get_kb_dir(company_name, kb_name), "excel")
    ensure_dir_exists(excel_dir)
    excel_path = os.path.join(excel_dir, f"{company_name}_{kb_name}_analysis.xlsx")
    wb.save(excel_path)
    
    logger.info(f"Excel report saved to {excel_path}")
    return excel_path


def generate_excel_report_from_kb(company_name, kb_name, prompt_type="prompt_1"):
    """Generate an Excel report from an existing knowledge base"""
    try:
        logger.info(f"Starting generate_excel_report_from_kb for {company_name}/{kb_name}")
        
        # Check if company_name and kb_name are valid
        if not company_name or not kb_name:
            logger.error("Company name or KB name is empty")
            return None
            
        # Get the retriever for this knowledge base
        logger.info(f"Getting retriever for {company_name}/{kb_name}")
        retriever = get_retriever_for_kb(company_name, kb_name)
        
        if not retriever:
            logger.error(f"Retriever is None - KB not found for {company_name}/{kb_name}")
            return None
            
        logger.info(f"Successfully got retriever, now generating Excel report...")
        
        # Generate the Excel report
        excel_path = generate_excel_report(company_name, kb_name, retriever, prompt_type=prompt_type)
        
        if not excel_path:
            logger.error("generate_excel_report returned None")
            return None
            
        # Return the path to the Excel file
        logger.info(f"Excel report generated successfully: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"Exception in generate_excel_report_from_kb: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
        
# def runquery(query_text, slide_index, shape_index, page_num, paragraph_index, retriever, ppt=None):
#     """Run a query and update a specific shape in the presentation"""
#     try:
#         # Get the response from the query
#         response = run_query(query_text, retriever)
        
#         # If no presentation is provided, just return the response
#         if ppt is None:
#             return response
        
#         # Get the slide
#         if slide_index < len(ppt.slides):
#             slide = ppt.slides[slide_index]
            
#             # Update the shape
#             if shape_index < len(slide.shapes):
#                 shape = slide.shapes[shape_index]
                
#                 if hasattr(shape, 'text_frame'):
#                     # If paragraph_index is specified, update only that paragraph
#                     if paragraph_index < len(shape.text_frame.paragraphs):
#                         shape.text_frame.paragraphs[paragraph_index].text = response
#                     else:
#                         shape.text_frame.text = response
                
                    
#                     return True
#         return False
#     except Exception as e:
#         logger.error(f"Error in runquery: {e}")
#         return False
    

def create_vector_store1(files, kb_name, company_name):
    """Create a vector store from uploaded files."""
    logger.info(f"Creating vector store for {company_name}/{kb_name}")
    
    try:
        # Sanitize KB name for collection name
        import re
        sanitized_kb_name = re.sub(r'[^a-zA-Z0-9_-]', '_', kb_name)
        if not sanitized_kb_name[0].isalnum():
            sanitized_kb_name = 'kb_' + sanitized_kb_name
        if len(sanitized_kb_name) > 0 and not sanitized_kb_name[-1].isalnum():
            sanitized_kb_name = sanitized_kb_name + '1'
        if len(sanitized_kb_name) < 3:
            sanitized_kb_name = sanitized_kb_name + '_db'
        sanitized_kb_name = sanitized_kb_name[:63]
        
        # Get directories
        kb_dir = get_kb_dir(company_name, kb_name)
        persist_directory = f"media/vector_stores/{company_name}/{sanitized_kb_name}"
        ensure_dir_exists(persist_directory)
        
        # Process uploaded files to get text
        all_texts = []
        file_paths = []
        
        # First save the files to disk
        for file_obj in files:
            # Save file to disk
            file_path = os.path.join(kb_dir, file_obj.name)
            with open(file_path, "wb") as f:
                for chunk in file_obj.chunks():
                    f.write(chunk)
            file_paths.append(file_path)
            logger.debug(f"Saved file: {file_path}")
        
        # Then process the files to extract text
        for file_path in file_paths:
            if file_path.lower().endswith('.pdf'):
                try:
                    loader = PyPDFLoader(file_path)
                    pages = loader.load_and_split()
                    all_texts.extend([page.page_content for page in pages])
                    logger.debug(f"Processed PDF: {file_path}")
                except Exception as e:
                    logger.error(f"Error processing PDF {file_path}: {e}")
        
        # Create chunks and vectorize
        chunks = text_splitter.create_documents(all_texts)
        uuids = [str(uuid4()) for _ in range(len(chunks))]
        
        # Define max batch size (lower than the 166 limit mentioned in the error)
        MAX_BATCH_SIZE = 100
        
        # Create batches of documents
        total_chunks = len(chunks)
        batches = []
        uuid_batches = []
        
        for i in range(0, total_chunks, MAX_BATCH_SIZE):
            end_idx = min(i + MAX_BATCH_SIZE, total_chunks)
            batches.append(chunks[i:end_idx])
            uuid_batches.append(uuids[i:end_idx])
        
        logger.info(f"Split {total_chunks} documents into {len(batches)} batches of maximum size {MAX_BATCH_SIZE}")
        
        # Create vector store
        vectorstore = Chroma(
            collection_name=sanitized_kb_name,
            embedding_function=embeddings,
            persist_directory=persist_directory,  # Fixed variable name
        )
        
        # Add documents in batches
        for i, (batch, uuid_batch) in enumerate(zip(batches, uuid_batches)):
            logger.info(f"Adding batch {i+1}/{len(batches)} with {len(batch)} documents")
            vectorstore.add_documents(documents=batch, ids=uuid_batch)
            
        logger.info(f"Successfully added all {total_chunks} document chunks to vector store")
        
        # Create and return a retriever
        retriever = vectorstore.as_retriever(search_type="similarity", search_kwargs={"k": 3})
        return retriever
        
    except Exception as e:
        logger.error(f"Error creating vector store: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None
    

def generate_chat_response(query, company_name, kb_name):
    """Generate a conversational response based on knowledge base content"""
    try:
        logger.info(f"Generating chat response for query: '{query}' using {company_name}/{kb_name}")

         # Check if KB name is None or empty
        if not kb_name:
            return "Please select a knowledge base before asking questions."
        
        
        
        # Check if KB name is None or empty
        if not kb_name:
            return "Please select a knowledge base before asking questions."
            
        # Check if company name is None or empty
        if not company_name:
            return "Please select a company before asking questions."
        
        
        
        
        # Load vector store and create retriever
        try:
            retriever = get_retriever_for_kb(company_name, kb_name)
        except Exception as e:
            logger.error(f"Error loading vector store: {e}")
            return "I'm having trouble accessing the knowledge base. Please try again later."
        
       
        
        # Create and run chain with timeout handling
        try:
            response = run_query(query, retriever)    
            return response
        except Exception as chain_error:
            return "I encountered an error while processing your request. Please try again later."
        
    except Exception as e:
        import traceback
        logger.error(f"Error generating chat response: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return "I encountered an error while processing your request. Please try again later."

def run_query(query, retriever, prompt_type="prompt_1"):
    """Run a query against the retriever and process with LLM"""
    if not llm:
        return "LLM not initialized. Cannot generate report content."
    
    # Get relevant documents
    retrieved_docs = retriever.invoke(query)
    context = "\n\n".join([doc.page_content for doc in retrieved_docs])
    
    selected_template_str = ""
    input_vars = ["context", "query"]
    prompt_data = {"context": context, "query": query}

    if prompt_type == "prompt_2":
        selected_template_str = PROMPT_2
        input_vars.append("length_guidance") # Add to input_vars
        prompt_data["length_guidance"] = "" # Pass empty string for this prompt type in run_query
        logger.info("Using Prompt 2 for run_query")
    else: # Default to prompt_1
        selected_template_str = PROMPT_1
        input_vars.append("length_guidance") # Add to input_vars
        prompt_data["length_guidance"] = ""
        logger.info("Using Prompt 1 for run_query")
        # input_vars is already ["context", "query"]

    prompt = PromptTemplate(
        template=selected_template_str,
        input_variables=input_vars
    )
    # # Create prompt template
    # prompt = PromptTemplate(
    #     template="""You are a professional business analyst creating a report.
        
    #     Based on the following context information, provide a well-structured, professional response to the query.
    #     Make your response comprehensive yet concise, suitable for a business presentation slide.
        
    #     Important formatting requirements:
    #     1. Start with a brief summary (1-2 sentences)
    #     2. Use bullet points for key information (• format)
    #     3. If providing data or statistics, present them clearly
    #     4. For any multi-part answers, use numbered lists
    #     5. Bold important terms or figures using markdown (**term**)
    #     6. Use short paragraphs with clear spacing between points
    #     7. Avoid long, dense paragraphs of text
        
    #     Context: {context}
        
    #     Query: {query}
        
    #     Response:""",
    #     input_variables=["context", "query"]
    # )
    
    # Create and run chain
    chain = LLMChain(llm=llm, prompt=prompt)
    response = chain.invoke(prompt_data)

    # Process response for better formatting in Excel
    text = response['text'].strip()

    # Convert markdown bullets to proper bullet points for Excel
    text = text.replace('\n• ', '\n• ')
    text = text.replace('\n* ', '\n• ')
    text = text.replace('\n- ', '\n• ')

    # Ensure proper line breaks between bullet points
    text = text.replace('\n\n• ', '\n• ')
    
    return text


def run_query_concise(query, retriever, prompt_type="prompt_1", max_length=None, analyst_answer_length=None):
    """
    Run a query against the retriever and process with LLM to produce concise answers
    similar in length to analyst answers
    """
    if not llm:
        return "LLM not initialized. Cannot generate report content."
    
    # Get relevant documents
    retrieved_docs = retriever.invoke(query)
    context = "\n\n".join([doc.page_content for doc in retrieved_docs])
    
    # Create length guidance based on analyst answer if provided
    current_length_guidance = ""
    if analyst_answer_length:
        current_length_guidance = f"Your answer should be approximately {analyst_answer_length} characters long."
    elif max_length:
        current_length_guidance = f"Your answer should not exceed {max_length} characters."
    else:
        current_length_guidance = "Keep your answer brief and concise."
    
    selected_template_str = ""
    # All prompts for run_query_concise will use context, query, and length_guidance
    input_vars = ["context", "query", "length_guidance"] 
    prompt_data = {"context": context, "query": query, "length_guidance": current_length_guidance}

    # # Create prompt template
    # prompt = PromptTemplate(
    #     template="""You are a professional business analyst creating a concise report.
        
    #     Based on the following context information, provide a brief, well-structured, professional response to the query.
    #     Make your response concise and to the point - include only the most relevant information.
        
    #     Important formatting and length requirements:
    #     1. {length_guidance}
    #     2. Use bullet points for listing key information (• format)
    #     3. Bold important terms or figures using markdown (**term**)
    #     4. Focus on the most important facts or insights only
    #     5. Avoid unnecessary context or background information
    #     6. Match the style and tone of an expert analyst's answer
        
    #     Context: {context}
        
    #     Query: {query}
        
    #     Response:""",
    #     input_variables=["context", "query", "length_guidance"]
    # )
    if prompt_type == "prompt_2":
        selected_template_str = PROMPT_2
        logger.info("Using Prompt 2 for run_query_concise")
    else: # Default to standard_qa_concise
        selected_template_str = PROMPT_1
        logger.info("Using Prompt 1 for run_query_concise")
        
    prompt = PromptTemplate(
        template=selected_template_str,
        input_variables=input_vars
    )

    # Create and run chain
    chain = LLMChain(llm=llm, prompt=prompt)
    response = chain.invoke(prompt_data)

    # Process response for better formatting in Excel
    text = response['text'].strip()

    # Convert markdown bullets to proper bullet points for Excel
    text = text.replace('\n• ', '\n• ')
    text = text.replace('\n* ', '\n• ')
    text = text.replace('\n- ', '\n• ')

    # Ensure proper line breaks between bullet points
    text = text.replace('\n\n• ', '\n• ')
    
    # return text
    return text, context



    
def get_excel_dir(company_name, kb_name):
    """Get Excel directory path and create if doesn't exist."""
    excel_path = os.path.join(get_kb_dir(company_name, kb_name), "excel")
    return ensure_dir_exists(excel_path)


def generate_qa_report_from_kb(company_name, kb_name, retriever,ground_truths_path=None, prompt_type="prompt_1"):
    """
    Generate a Q&A report for the specified company and knowledge base

    Args:
        company_name: Name of the company
        kb_name: Name of the knowledge base
        retriever: The retriever object for the knowledge base
        ground_truths_path: Optional path to ground truths file for RAGAS evaluation
    
    Returns:
        Path to the generated Excel report
    """
    # Define the predefined questions and categories
    qa_data = {
        "Fact Extraction": [
            {
                "question": f"What is the full name of the {company_name}?",
                "analyst_answer": "ADVANCED MICRO DEVICES, INC."
            },
            {
                "question": f"Which all {company_name} documents do you have access to?",
                "analyst_answer": "10K reports of AMD from 2014 till 2023"
            },
            {
                "question": f"What is the address of{company_name}'s headquarters?",
                "analyst_answer": "2485 Augustine Drive, Santa Clara, California 95054, United States"
            },
            {
                "question": f"Who are the key management team members of {company_name} as of the latest filing?",
                "analyst_answer": """Below are the key members of the management team:
                * President and CEO, Director: Lisa T. Su
                * Executive Vice President, Chief Financial Officer and  
                  Treasurer: Jean Hu
                * Corporate Vice President, Chief Accounting Officer: Darla Smith"""
            },
            {
                "question": f"What is the total number of patents that {company_name} has as of the latest reporting?",
                "analyst_answer": "As per the company's 10K FY2023: they had approximately 7,500 patents in the United States and approximately 2,000 patent applications pending in the United States; Including United States and foreign matters, they have approximately 18,500 patent matters worldwide consisting of approximately 12,800 issued patents and 5,600 patent applications pending."
            },
        ],
        "Summarization": [
            {
                "question": f"Give a brief history of {company_name}.",
                "analyst_answer": "AMD, a global semiconductor company, was incorporated in 1969 as a Silicon Valley start-up with dozens of employees focused on leading-edge semiconductor products, and became public in 1972. Today, they have grown into a global company achieving many important industry firsts along the way. They develop high-performance and adaptive computing to solve some of the world’s toughest and most interesting challenges."
            },
            {
                "question": f"What are the main products of {company_name}? Give a brief description.",
                "analyst_answer": "AMD’s products include x86 microprocessors (CPUs) and graphics processing units (GPUs), as standalone devices or as incorporated into accelerated processing units (APUs), chipsets, data center and professional GPUs, embedded processors, semi-custom System-on-Chip(SoC) products, microprocessor and SoC development services and technology, data processing units (DPUs), Field Programmable Gate Arrays (FPGAs),System on Modules (SOMs), Smart Network Interface Cards (SmartNICs), AI Accelerators and Adaptive SoC products."
            },
            {
                "question": f"What are the main revenue segments of {company_name} at the end of 2023?",
                "analyst_answer": """Major revenue segments of AMD:
                * Data Center: 29% of net revenue
                * Client: 21% of net revenue
                * Gaming: 27% of net revenue
                * Embedded: 23% of net revenue"""
            },
            {
                "question": f"Who are the main competitors of {company_name} as of 2023?",
                "analyst_answer": """Segment wise competitors:
                * Data Center: Nvidia and Intel
                * Client Segment: Intel
                * Gaming Segment: Nvidia, Intel
                * Embedded Segment: Intel, Lattice Semiconductor and 
                  Microsemi Corporation (Microsemi,acquired by Microchip), from ASSP vendors such as Broadcom Corporation, Marvell Technology Group, Analog Devices, Texas Instruments and NXP Semiconductors, and from NVIDIA"""
            },
            {
                "question": f"What are the major risk factors for {company_name}?",
                "analyst_answer": """
                * Intel Corporation’s dominance of the microprocessor market and its aggressive business practices may limit AMD's ability to compete effectively on a level playing field
                * Cyclicity of the semiconductor industry, and the fluctuation of demand for products
                * Success for AMD is dependent upon its ability to introduce products on a timely basis with features and performance levels that provide value to their customers while supporting and coinciding with significant industry transitions; so consistent innovation and product upgradation is required
                * AMD relies on third parties to manufacture its products, and if they are unable to do so on a timely basis in sufficient quantities and using competitive technologies, AMD's business could be materially adversely affected
                * If AMD loses Microsoft Corporation’s support for their products or other software vendors do not design and develop software to run on their products, their ability to sell their products could be materially adversely affected
                * Government actions and regulations such as export regulations, tariffs, and trade protection measures may limit AMD's ability to export our products to certain customers"""
            },
            {
                "question": f"What are the acquisitions {company_name} has done over the last 5 years?",
                "analyst_answer": """
                * October 2023, they acquired Nod Inc, an open AI software company
                * August 2023, AMD acquired Mipsology SAS, an AI software company
                * May 2022, AMD acquired Pensando for, a next-generation distributed computing company
                * February 2022, AMD acquired Xilinx, a provider of adaptive computing solutions"""
            }
        ],
        "Analysis": [
            {
                "question": f"Is {company_name} operating in a crowded market?",
                "analyst_answer": "AMD operates in a highly competitive market, and the company expects the competition to continue to be intense due to rapid technological changes, frequent product introductions by its competitors or new competitors of products that may provide better performance or experience or that may include additional features that render AMD's products comparatively less competitive"
            },
            {
                "question": f"What is {company_name}'s market share in its major revenue segment?",
                "analyst_answer": """Industry size/market size data is not available, however, as for AMD's net revenue for financial year 2023, below is how its segments contributed to its total revenue of $22,680 Mn:

                * Data Center: 29%
                * Client: 21%
                * Gaming: 27%
                * Embedded: 23%"""
            },
            {
                "question": f"How is the corporate governance at {company_name} with respect to disclosures, independent directors as in 2023?",
                "analyst_answer": "Corporate Governance Principles of the company are posted on an external link, and its not available in the 10K reports. Here is the link for the same: https://www.amd.com/en/corporate/corporate-responsibility.html"
            }
        ]
    }
    
    try:
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"{company_name} Q&A Analysis"
        
        # Define styles
        header_font = Font(bold=True, size=12)
        category_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="CFCFCF", end_color="CFCFCF", fill_type="solid")
        category_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add report title
        # ws.merge_cells('A1:D1')
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{company_name} - Question & Answer Analysis Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add headers
        ws['A3'] = "Question Type"
        ws['B3'] = "Question"
        ws['C3'] = "Analyst's Answer"
        ws['D3'] = "AI Agent Answer"
        ws['E3'] = "Retrieved Context"
        
        for cell in ['A3', 'B3', 'C3', 'D3', 'E3']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40

        # Starting row for data
        row = 4
        
        # Iterate through categories and questions
        for category, questions in qa_data.items():
            # Add category row
            ws.merge_cells(f'A{row}:H{row}')
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = category_font
            ws[f'A{row}'].fill = category_fill
            ws[f'A{row}'].alignment = Alignment(horizontal='left')
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].border = thin_border
            row += 1
            
            # Process each question in the category
            for q_data in questions:
                question = q_data["question"]
                analyst_answer = q_data["analyst_answer"]


                # Calculate the length of the analyst's answer to guide AI response length
                analyst_answer_length = len(analyst_answer)
                
                # Get AI answer from the retriever and LLM with length guidance
                ai_answer, retrieved_context = run_query_concise(
                    question, 
                    retriever, 
                    analyst_answer_length=analyst_answer_length,
                    prompt_type=prompt_type
                )
                
                # Add to worksheet
                ws[f'A{row}'] = category
                ws[f'B{row}'] = question
                ws[f'C{row}'] = analyst_answer
                ws[f'D{row}'] = ai_answer
                ws[f'E{row}'] = retrieved_context
                
                # Apply borders and text wrapping
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].border = thin_border
                    ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                
                row += 1
        
        # Save the workbook
        excel_dir = os.path.join(get_kb_dir(company_name, kb_name), "excel")
        ensure_dir_exists(excel_dir)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = os.path.join(excel_dir, f"{company_name}_{kb_name}_qa_analysis_{timestamp}.xlsx")
        wb.save(excel_path)
        
        logger.info(f"Q&A Excel report saved to {excel_path}")

         # If ground truths file is provided, run RAGAS evaluation
        if ground_truths_path and os.path.exists(ground_truths_path):
            logger.info(f"Ground truths file found at {ground_truths_path}, running RAGAS evaluation")
            # Pass company_name and kb_name to use for retriever access
            evaluated_path = evaluate_qa_report_with_ragas(
                excel_path, 
                ground_truths_path,
                company_name,
                kb_name,
            )
            if evaluated_path:
                logger.info(f"RAGAS evaluation completed, returning evaluated report path")
                return evaluated_path

            else:
                logger.error(f"RAGAS evaluation failed, returning original excel path")
        else:
            if ground_truths_path:
                logger.error(f"Ground truths file not found at {ground_truths_path}")
            else:
                logger.info("No ground truths file provided, skipping RAGAS evaluation")
        return excel_path
        
    except Exception as e:
        logger.error(f"Error generating Q&A report: {e}")
        logger.error(traceback.format_exc())
        return None
    
# Update get_retriever_for_kb function
def get_retriever_for_kb(company_name, kb_name):
    """
    Get a retriever for an existing knowledge base
    """
    try:
        logger.info(f"Getting retriever for {company_name}/{kb_name}")
        # Use the same embeddings as the rest of your application
        from langchain_community.embeddings import SentenceTransformerEmbeddings
        
        # Sanitize KB name for use as collection name (same as in create_vector_store1)
        import re
        sanitized_kb_name = re.sub(r'[^a-zA-Z0-9_-]', '_', kb_name)
        if not sanitized_kb_name[0].isalnum():
            sanitized_kb_name = 'kb_' + sanitized_kb_name
        if len(sanitized_kb_name) > 0 and not sanitized_kb_name[-1].isalnum():
            sanitized_kb_name = sanitized_kb_name + '1'
        if len(sanitized_kb_name) < 3:
            sanitized_kb_name = sanitized_kb_name + '_db'
        sanitized_kb_name = sanitized_kb_name[:63]

        # Path to the vector store
        persist_directory = f"media/vector_stores/{company_name}/{sanitized_kb_name}"
        
        # Check if directory exists
        if not os.path.exists(persist_directory):
            logger.error(f"Vector store directory not found: {persist_directory}")
            return None
        
        # Initialize embeddings - USE THE SAME EMBEDDINGS AS IN create_vector_store1
        embeddings_model = SentenceTransformerEmbeddings(model_name="all-MiniLM-L6-v2")
        
        # Load the vector store
        vectordb = Chroma(
            persist_directory=persist_directory,
            embedding_function=embeddings_model,
            collection_name=sanitized_kb_name
        )
        
        # Create retriever
        retriever = vectordb.as_retriever(
            search_type="similarity",
            search_kwargs={"k": 5}
        )
        
        logger.info(f"Successfully loaded retriever for {company_name}/{kb_name}")
        return retriever
        
    except Exception as e:
        import traceback
        logger.error(f"Error loading retriever for {company_name}/{kb_name}: {e}")
        logger.error(traceback.format_exc())
        return None
    

def evaluate_qa_report_with_ragas(excel_path, ground_truths_path, company_name, kb_name):
    """
    Evaluate the QA report using RAGAS metrics and update the Excel file
    
    Args:
        excel_path: Path to the generated Excel report
        ground_truths_path: Path to the ground_truths file (.md or .txt)
        company_name: Name of the company for retriever access
        kb_name: Name of the knowledge base for retriever access
    
    Returns:
        Path to the updated Excel file with evaluation metrics
    """

       

    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import re
        import ragas
        from ragas import evaluate
        logger.info(f"RAGAS version: {ragas.__version__}")
        logger.info(f"Starting RAGAS evaluation for report: {excel_path}")


        # Add this function to report_generator1.py
        def normalize_question(text):
            """Normalize a question text for better matching"""
            if not text:
                return ""
            # Remove punctuation
            text = re.sub(r'[^\w\s]', '', text)
            # Lowercase
            text = text.lower()
            # Remove extra whitespace
            text = re.sub(r'\s+', ' ', text).strip()
            return text
        
         # Test retriever with a simple query to validate it's working
        try:
            # Get retriever for this knowledge base
            retriever = get_retriever_for_kb(company_name, kb_name)
            if not retriever:
                logger.error(f"Could not get retriever for {company_name}/{kb_name}")
                return None
                
            test_docs = retriever.invoke("What is AMD?")
            if test_docs:
                logger.info(f"Retriever test successful. Got {len(test_docs)} documents.")
                logger.info(f"Sample content: {test_docs[0].page_content[:100]}...")
            else:
                logger.warning("Retriever test returned no documents!")
        except Exception as e:
            logger.error(f"Retriever test failed: {e}")
            return None

        
        # Initialize LLM and embedding wrappers for RAGAS
        try:
            # Import RAGAS dependencies
            from ragas.llms import LangchainLLMWrapper
            from ragas.embeddings import LangchainEmbeddingsWrapper
            from ragas.dataset_schema import SingleTurnSample 
            from ragas.metrics import (
                LLMContextPrecisionWithReference,
                # LLMContextRecall,
                # ContextEntityRecall,
                # NoiseSensitivity,
                answer_relevancy,
                Faithfulness,
                answer_correctness
            )
            from langchain_core.documents import Document
            
            # Define LLM wrapper for metrics
            ragas_llm = LangchainLLMWrapper(llm)
            
            # Define embeddings wrapper for metrics that need it
            ragas_embeddings = LangchainEmbeddingsWrapper(embeddings)
            
            # Define metrics with proper wrappers
            metrics = {
                'E': ('Faithfulness', Faithfulness(llm=ragas_llm)),
                'F': ('Response Relevancy', answer_relevancy),
                'G': ('Context Precision', LLMContextPrecisionWithReference(llm=ragas_llm)),
                'H': ('Factual Correctness', answer_correctness)                # 'H': ('Context Recall', LLMContextRecall(llm=ragas_llm)),
                # 'I': ('Entity Recall', ContextEntityRecall(llm=ragas_llm)),
                # 'J': ('Noise Sensitivity', NoiseSensitivity(llm=ragas_llm))
            }
            
            logger.info("Successfully initialized RAGAS metrics")

            # try:
            #     from ragas.metrics import answer_relevancy
            #     metrics['F'] = ('Relevancy', answer_relevancy)
            #     logger.info("✓ Added Relevancy metric")
            # except ImportError:
            #     logger.warning("Could not import answer_relevancy metric")
            
        except ImportError as ie:
            logger.error(f"RAGAS library import error: {ie}")
            logger.error(traceback.format_exc())
            return None
        except Exception as e:
            logger.error(f"Error initializing RAGAS metrics: {e}")
            logger.error(traceback.format_exc())
            return None
        
        # Get retriever for this knowledge base
        retriever = get_retriever_for_kb(company_name, kb_name)
        if not retriever:
            logger.error(f"Could not get retriever for {company_name}/{kb_name}")
            return None
        
        # Load the ground truths from the file
        with open(ground_truths_path, 'r', encoding='utf-8') as f:
            ground_truths_content = f.read()

        logger.info(f"Ground truths file contents preview: {ground_truths_content[:200]}...")
        logger.info(f"File size: {len(ground_truths_content)} characters")
        
        # Parse the ground truths file based on its format
        ground_truths = []
        file_extension = os.path.splitext(ground_truths_path.lower())[1]
        
        if file_extension == '.md':
            # Parse for questions and ground truths
            q_matches = re.findall(r'## Question \d+:(.*?)(?=## Ground Truth:)', ground_truths_content, re.DOTALL)
            gt_matches = re.findall(r'## Ground Truth:(.*?)(?=## Question \d+:|$)', ground_truths_content, re.DOTALL)

            if len(q_matches) == len(gt_matches):
                for i in range(len(q_matches)):
                    ground_truths.append({
                        "question": q_matches[i].strip(),
                        "ground_truth": gt_matches[i].strip()
                    })
        
        elif file_extension == '.txt':
            # Parse TXT format with QUESTION: and GROUND TRUTH: markers
            entries = ground_truths_content.split('---')
            
            for entry in entries:
                entry = entry.strip()
                if not entry:
                    continue
                
                question_match = re.search(r'QUESTION:(.*?)(?=GROUND TRUTH:|$)', entry, re.DOTALL, re.IGNORECASE)
                ground_truth_match = re.search(r'GROUND TRUTH:(.*?)(?=CONTEXT:|$)', entry, re.DOTALL, re.IGNORECASE)
                
                if question_match and ground_truth_match:
                    question = question_match.group(1).strip()
                    ground_truth = ground_truth_match.group(1).strip()
                    ground_truths.append({
                        "question": question,
                        "ground_truth": ground_truth
                    })
        
        if not ground_truths:
            logger.error(f"Could not parse any questions from ground truths file: {ground_truths_path}")
            logger.error(f"File content preview: {ground_truths_content[:500]}...")
            return None
            
        logger.info(f"Parsed {len(ground_truths)} questions from ground truths file")
        
        # For each question, retrieve context using the retriever
        # Replace the context retrieval section in evaluate_qa_report_with_ragas
        for q_data in ground_truths:
            question = q_data["question"]
            
            # Use the retriever to get relevant context
            try:
                # Add debug log
                logger.info(f"Retrieving context for ground truth question: '{question}'")
                
                # Test retriever first
                retrieved_docs = retriever.invoke(question)
                
                if retrieved_docs and len(retrieved_docs) > 0:
                    # Log success
                    logger.info(f"✓ Retrieved {len(retrieved_docs)} documents for ground truth question")
                    
                    # Get text from docs and join
                    context_texts = [doc.page_content for doc in retrieved_docs if hasattr(doc, 'page_content') and doc.page_content]
                    context = "\n\n".join(context_texts)
                    q_data["context"] = context
                    
                    # Convert to Document objects explicitly
                    q_data["context_docs"] = []
                    for doc in retrieved_docs:
                        if hasattr(doc, 'page_content') and doc.page_content:
                            q_data["context_docs"].append(
                                Document(page_content=doc.page_content, 
                                        metadata=doc.metadata if hasattr(doc, 'metadata') else {})
                            )
                    
                    # Verify we have context docs
                    logger.info(f"✓ Created {len(q_data['context_docs'])} context documents")
                    if len(q_data["context_docs"]) > 0:
                        logger.info(f"  First context doc preview: {q_data['context_docs'][0].page_content[:100]}...")
                else:
                    logger.warning(f"✗ No context found for ground truth question: '{question}'")
                    # For debugging: try a simpler query
                    test_docs = retriever.invoke("AMD")
                    if test_docs:
                        logger.info(f"Test query 'AMD' returned {len(test_docs)} docs, but original question returned none")
                    
                    # Still create empty context to avoid errors
                    q_data["context"] = ""
                    q_data["context_docs"] = []
            except Exception as e:
                logger.error(f"Error retrieving context for ground truth question '{question}': {e}")
                logger.error(traceback.format_exc())
                q_data["context"] = ""
                q_data["context_docs"] = []
        
        # Load the Excel file
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Add headers for evaluation metrics
        for col, (header_name, _) in metrics.items():
            ws[f'{col}3'] = header_name
        
        # Apply header styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CFCFCF", end_color="CFCFCF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in metrics.keys():
            ws[f'{col}3'].font = header_font
            ws[f'{col}3'].fill = header_fill
            ws[f'{col}3'].border = thin_border
            ws.column_dimensions[col].width = 15
        
        # Find and evaluate each question
        row = 5  # Starting from the first question (after headers and category)
        metrics_by_category = {col: {} for col in metrics.keys()}
        
        # # Define a function to evaluate all metrics for a single QA pair
        # def evaluate_metrics(question, answer, contexts, ground_truth, metrics_dict):
        #     results = {}
            
        #     logger.info(f"Evaluating: '{question[:50]}...'")
        #     logger.info(f"  Answer: '{answer[:50]}...'")
        #     logger.info(f"  Ground truth: '{ground_truth[:50]}...'")
        #     logger.info(f"  Contexts: {len(contexts)} documents")
            
        #     # try:
        #     #     # Try different import paths based on RAGAS version
        #     #     try:
        #     #         from ragas.single_turn import SingleTurnSample
        #     #         logger.info("Using ragas.single_turn.SingleTurnSample")
        #     #     except ImportError:
        #     #         try:
        #     #             # For older versions
        #     #             from ragas.evaluation import SingleTurnSample
        #     #             logger.info("Using ragas.evaluation.SingleTurnSample")
        #     #         except ImportError:
        #     #             # For newer versions
        #     #             from ragas import evaluate_ragas_sample
        #     #             logger.info("Using ragas.evaluate_ragas_sample approach")
                        
        #     #             # In this case, we'll use evaluate_ragas_sample directly instead of SingleTurnSample
        #     #             for col, (metric_name, metric) in metrics_dict.items():
        #     #                 try:
        #     #                     # Create a dict for evaluate_ragas_sample
        #     #                     sample = {
        #     #                         "question": question,
        #     #                         "answer": answer,
        #     #                         "contexts": [doc.page_content for doc in contexts],
        #     #                         "ground_truth": ground_truth
        #     #                     }
        #     #                     score = metric.score(sample)
        #     #                     logger.info(f"{metric_name} score: {score}")
        #     #                     results[col] = score
        #     #                 except Exception as e:
        #     #                     logger.error(f"Error calculating {metric_name}: {e}")
        #     #                     logger.error(traceback.format_exc())
        #     #                     results[col] = None
        #     #             return results
                
        #     #     # Create RAGAS sample the standard way if imports worked
        #     #     sample = SingleTurnSample(
        #     #         question=question,
        #     #         answer=answer,
        #     #         contexts=contexts,
        #     #         ground_truth=ground_truth
        #     #     )
                
        #     #     # Evaluate each metric
        #     #     for col, (metric_name, metric) in metrics_dict.items():
        #     #         try:
        #     #             logger.info(f"Calculating {metric_name}...")
        #     #             score = metric.single_turn_score(sample)
        #     #             logger.info(f"{metric_name} score: {score}")
        #     #             results[col] = score
        #     #         except Exception as e:
        #     #             logger.error(f"Error calculating {metric_name}: {e}")
        #     #             logger.error(traceback.format_exc())
        #     #             results[col] = None
        #     # except Exception as e:
        #     #     logger.error(f"Error evaluating metrics: {e}")
        #     #     logger.error(traceback.format_exc())
            
        #     # return results        

        #     for col, (metric_name, metric_func) in metrics_dict.items():
        #         try:
        #             logger.info(f"  Calculating {metric_name}...")
                    
        #             # For faithfulness metric
        #             if metric_name == 'Faithfulness':
        #                 # sample = SingleTurnSample(
        #                 #     user_input=question,
        #                 #     response=answer,
        #                 #     retrieved_contexts=[doc.page_content for doc in contexts]
        #                 # )
        #                 # value = metric_func.single_turn_ascore(sample)
        #                 # print(value)
        #                 # score_df = metric_func.score(
        #                 #     question=[question],
        #                 #     answer=[answer],
        #                 #     contexts=[[doc for doc in contexts]]
        #                 # )
        #                 # value = score_df['faithfulness'].iloc[0]
                        
        #             # For relevancy metric
        #             elif metric_name == 'Relevancy':
        #                 print(-1)
        #                 # result = evaluate(
        #                 #     dataset=Dataset.from_dict({
        #                 #         "question": [question],
        #                 #         "answer": [answer],
        #                 #         "contexts": [[doc.page_content for doc in contexts]],
        #                 #         "ground_truth": [ground_truth]
        #                 #     }),
        #                 #     metrics=[metric_func]
        #                 # )
        #                 # value = result[metric_name][0]
        #             else:
        #                 logger.warning(f"Unsupported metric: {metric_name}")
        #                 continue
                        
        #             logger.info(f"  ✓ {metric_name} score: {value:.4f}")
        #             results[col] = value
                    
        #         except Exception as e:
        #             logger.error(f"  ✗ Error calculating {metric_name}: {e}")
        #             logger.error(traceback.format_exc())
        #             results[col] = None
                    
        #     return results

        def evaluate_metrics(question, answer, contexts, ground_truth, metrics_dict, ragas_llm_local, ragas_embeddings_local):
            """
            Evaluates a single QA pair against a dictionary of RAGAS metrics.

            Args:
                question (str): The question.
                answer (str): The AI-generated answer.
                contexts (list): A list of langchain_core.documents.base.Document objects.
                ground_truth (str): The reference/analyst answer.
                metrics_dict (dict): A dictionary where keys are Excel column letters and
                                    values are tuples of (display_name, metric_object).
                ragas_llm_local: The LangchainLLMWrapper instance for RAGAS.
                ragas_embeddings_local: The LangchainEmbeddingsWrapper instance for RAGAS.

            Returns:
                dict: A dictionary mapping Excel column letters to calculated scores.
            """
            results_scores = {}
            
            # Log input types for sanity check (can be removed after confirming)
            logger.debug(f"evaluate_metrics - question type: {type(question)}")
            logger.debug(f"evaluate_metrics - answer type: {type(answer)}")
            logger.debug(f"evaluate_metrics - contexts type: {type(contexts)}")
            if contexts:
                logger.debug(f"evaluate_metrics - contexts[0] type: {type(contexts[0])}")
            logger.debug(f"evaluate_metrics - ground_truth type: {type(ground_truth)}")

            for col_letter, (metric_display_name, metric_object) in metrics_dict.items():
                try:
                    logger.info(f"  Calculating {metric_display_name} for question: '{question[:50]}...'")

                    # Prepare the data in the format RAGAS evaluate expects for a single item.
                    # 'question', 'answer', 'ground_truth' are lists of strings.
                    # 'contexts' is a list of lists of strings.
                    current_eval_data = {
                        "question": [question],
                        "answer": [answer],
                        "contexts": [[doc.page_content for doc in contexts if hasattr(doc, 'page_content') and doc.page_content]],
                        "ground_truth": [ground_truth]
                    }
                    
                    # Ensure contexts list is not empty if there are no valid page_contents
                    if not current_eval_data["contexts"][0] and contexts: # if original contexts list was not empty but produced no content
                        logger.warning(f"  Context list became empty after extracting page_content for {metric_display_name}. Using empty list for contexts.")
                        current_eval_data["contexts"] = [[]] # RAGAS expects list of lists, even if inner list is empty
                    elif not contexts: # If original contexts list was empty
                        current_eval_data["contexts"] = [[]]


                    # Create a Dataset object for the current sample
                    # This import should ideally be at the top of the file
                    from datasets import Dataset
                    current_dataset = Dataset.from_dict(current_eval_data)

                    # Evaluate using the specific metric object
                    evaluation_result_obj = evaluate(
                        dataset=current_dataset,
                        metrics=[metric_object],  # Pass the specific metric instance
                        llm=ragas_llm_local,        # Pass the LLM wrapper
                        embeddings=ragas_embeddings_local, # Pass the embeddings wrapper
                        raise_exceptions=True # Set to True to catch specific errors during metric calculation
                    )
                    
                    # The result of evaluate is an EvaluationResult object.
                    # It behaves like a dictionary where keys are metric names (metric_object.name).
                    metric_internal_name = metric_object.name 
                    
                    # Access the score
                    # For a single-row dataset and single metric, the value should be a float.
                    score_value = evaluation_result_obj[metric_internal_name]

                    # RAGAS sometimes returns the score within a list if it processes it as a batch of 1.
                    if isinstance(score_value, list) and len(score_value) == 1:
                        final_score = float(score_value[0])
                    elif isinstance(score_value, (float, int)):
                        final_score = float(score_value)
                    else:
                        logger.error(f"  Unexpected score format for {metric_display_name}: {score_value} (type: {type(score_value)})")
                        final_score = None # Or handle as an error, e.g., by setting a specific error string

                    if final_score is not None:
                        logger.info(f"  ✓ {metric_display_name} score: {final_score:.4f}")
                        results_scores[col_letter] = final_score
                    else:
                        # This path is taken if final_score couldn't be determined
                        logger.warning(f"  Score for {metric_display_name} is None or in an unexpected format.")
                        results_scores[col_letter] = "Err" 
                        
                except Exception as e:
                    logger.error(f"  ✗ Error calculating {metric_display_name} for question '{question[:50]}...': {e}")
                    import traceback
                    logger.error(traceback.format_exc())
                    results_scores[col_letter] = "Error" # Mark as Error in Excel
                    
            return results_scores

        while row <= ws.max_row:
            # Skip if row is a category header (merged cells)
            try:
                if ws.cell(row=row, column=1).value == ws.cell(row=row, column=2).value:
                    row += 1
                    continue
            except:
                # If error checking merged cells, just continue
                pass
                
            # Get question, category, and AI answer from excel
            question_cell_value = ws.cell(row=row, column=2).value
            if not question_cell_value:
                row += 1
                continue
                
            category = ws.cell(row=row, column=1).value
            ai_answer = ws.cell(row=row, column=4).value
            
            if not ai_answer:
                row += 1
                continue
            
            # Find matching ground truth
            matching_gt = None
            best_match_score = 0

            # First make sure we have question_cell_value
            if not question_cell_value:
                logger.warning("Question cell value is empty, skipping this row")
                row += 1
                continue

            question_norm = normalize_question(question_cell_value)


            # Log the current question we're trying to match
            logger.info(f"Trying to match question: '{question_cell_value}'")
            logger.info(f"Normalized question: '{question_norm}'")

            # Debug all available ground truth questions
            logger.info(f"Available ground truth questions ({len(ground_truths)}):")
            for idx, gt in enumerate(ground_truths):
                gt_norm = normalize_question(gt["question"])
                logger.info(f"  GT #{idx+1}: '{gt['question']}'")
                logger.info(f"  GT #{idx+1} normalized: '{gt_norm}'")
                
                # Check for exact match (normalized)
                if gt_norm == question_norm:
                    matching_gt = gt
                    logger.info(f"  EXACT MATCH FOUND!")
                    break
                
                # Calculate similarity score
                # Jaccard similarity on word sets
                q1_words = set(gt_norm.split())
                q2_words = set(question_norm.split())
                
                if q1_words and q2_words:
                    # Calculate Jaccard similarity
                    similarity = len(q1_words.intersection(q2_words)) / len(q1_words.union(q2_words))
                    logger.info(f"  Similarity: {similarity:.2f}")
                    
                    if similarity > 0.5 and similarity > best_match_score:  # At least 50% similar
                        matching_gt = gt
                        best_match_score = similarity
                        logger.info(f"  NEW BEST MATCH: Score {similarity:.2f}")

            if matching_gt:
                logger.info(f"Found matching ground truth: '{matching_gt['question']}'")
            else:
                logger.info(f"No matching ground truth found for: '{question_cell_value}'")
                
                # Try to find any match if nothing found with good similarity
                if not matching_gt and ground_truths:
                    # If no match found but we have ground truths, use the first one as a fallback
                    logger.warning(f"No good match found, trying with lower threshold")
                    
                    # Simple word overlap scoring
                    best_match = None
                    best_score = 0
                    
                    # 1. First try exact match
                    for gt in ground_truths:
                        if normalize_question(gt["question"]) == question_norm:
                            matching_gt = gt
                            logger.info(f"✓ EXACT MATCH: '{gt['question']}'")
                            break

                    # 2. If no exact match, try substring matching
                    if not matching_gt:
                        for gt in ground_truths:
                            gt_norm = normalize_question(gt["question"])
                            # Check if one is substring of the other
                            if gt_norm in question_norm or question_norm in gt_norm:
                                matching_gt = gt
                                logger.info(f"✓ SUBSTRING MATCH: '{gt['question']}'")
                                break
                    # 3. If still no match, use word overlap similarity
                    if not matching_gt:
                        for gt in ground_truths:
                            gt_norm = normalize_question(gt["question"])
                            # Calculate word overlap
                            q1_words = set(gt_norm.split())
                            q2_words = set(question_norm.split())
                            
                            if q1_words and q2_words:
                                # Calculate Jaccard similarity - just once per ground truth
                                intersection = len(q1_words.intersection(q2_words))
                                union = len(q1_words.union(q2_words))
                                similarity = intersection / union if union > 0 else 0
                                
                                if similarity > 0.5 and similarity > best_match_score:
                                    matching_gt = gt
                                    best_match_score = similarity
                                    logger.info(f"✓ SIMILARITY MATCH ({similarity:.2f}): '{gt['question']}'")
                    # Log the final result
                    if matching_gt:
                        logger.info(f"✓ MATCH FOUND: '{matching_gt['question']}'")
                    else:
                        logger.info(f"✗ NO MATCH FOUND for: '{question_cell_value}'")
                                        
            if matching_gt and "context_docs" in matching_gt and matching_gt["context_docs"]:
                # Calculate RAGAS metrics
                try:
                    # Create explicit parameters
                    eval_question = question_cell_value
                    eval_answer = ai_answer
                    eval_contexts = matching_gt["context_docs"]
                    eval_ground_truth = matching_gt["ground_truth"]
                    
                    logger.info(f"Evaluating question with {len(eval_contexts)} context documents")
                    # Get all scores at once
                    # metric_results = evaluate_metrics(
                    #     question=question_cell_value,
                    #     answer=ai_answer,
                    #     contexts=matching_gt["context_docs"],
                    #     ground_truth=matching_gt["ground_truth"],
                    #     metrics_dict=metrics
                    # )
                    metric_results = evaluate_metrics(
                        question=question_cell_value,       # This is eval_question
                        answer=ai_answer,                   # This is eval_answer
                        contexts=matching_gt["context_docs"], # This is eval_contexts
                        ground_truth=matching_gt["ground_truth"], # This is eval_ground_truth
                        metrics_dict=metrics,
                        ragas_llm_local=ragas_llm,          # Pass the wrapper
                        ragas_embeddings_local=ragas_embeddings # Pass the wrapper
                    )
                    # Add scores to Excel and tracking
                    for col, score in metric_results.items():
                        if score is not None:
                            # Initialize category tracking if needed
                            if category not in metrics_by_category[col]:
                                metrics_by_category[col][category] = []
                            
                            # Add to Excel
                            ws.cell(row=row,column=ord(col) - ord('A') + 1).value = round(score, 2)
                            
                            # Track for averages
                            metrics_by_category[col][category].append(score)
                        else:
                            ws.cell(row=row,column=ord(col) - ord('A') + 1).value = "Error"

                except Exception as eval_error:
                    logger.error(f"Error evaluating question '{question_cell_value}': {eval_error}")
                    logger.error(traceback.format_exc())
                    # Add N/A for failed evaluations
                    for col in metrics.keys():
                        ws.cell(row=row, column=ord(col) - ord('A') + 1).value = "N/A"
            else:
                if not matching_gt:
                    logger.error(f"No matching ground truth found for: '{question_cell_value}'")
                elif not "context_docs" in matching_gt:
                    logger.error(f"Ground truth match found but no 'context_docs' key: {matching_gt.keys()}")
                elif not matching_gt["context_docs"]:
                    logger.error(f"Ground truth match found but 'context_docs' is empty")
                else:
                    logger.info(f"Ground truth and context found! Ready to evaluate.")
                # No matching ground truth or no context found
                for col in metrics.keys():
                    ws.cell(row=row, column=ord(col) - ord('A') + 1).value = "No GT/Context"
            row +=1
                                
        # Add a summary section
        row = ws.max_row + 2
        max_col = chr(ord('A') + len(metrics))
        ws.merge_cells(f'A{row}:{max_col}{row}')
        ws[f'A{row}'] = "RAGAS Evaluation Summary"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        # Add headers for summary
        row += 2
        ws[f'A{row}'] = "Category"
        
        col_index = 'B'
        for col, (metric_name, _) in metrics.items():
            ws[f'{col_index}{row}'] = metric_name
            ws[f'{col_index}{row}'].font = Font(bold=True)
            ws[f'{col_index}{row}'].fill = PatternFill(start_color="CFCFCF", end_color="CFCFCF", fill_type="solid")
            col_index = chr(ord(col_index) + 1)
        
        # Add category averages
        categories = set()
        for col_data in metrics_by_category.values():
            categories.update(col_data.keys())
        
        for category in categories:
            row += 1
            ws[f'A{row}'] = category
            
            col_index = 'B'
            for col in metrics.keys():
                if category in metrics_by_category[col] and metrics_by_category[col][category]:
                    values = metrics_by_category[col][category]
                    ws[f'{col_index}{row}'] = round(sum(values) / len(values), 2)
                col_index = chr(ord(col_index) + 1)
        
        # Add overall average
        row += 2
        ws[f'A{row}'] = "OVERALL AVERAGE"
        ws[f'A{row}'].font = Font(bold=True)
        
        # Calculate overall averages
        col_index = 'B'
        for col in metrics.keys():
            all_values = []
            for category_values in metrics_by_category[col].values():
                all_values.extend(category_values)
                
            if all_values:
                ws[f'{col_index}{row}'] = round(sum(all_values) / len(all_values), 2)
            col_index = chr(ord(col_index) + 1)
        
        # Add color coding to the scores
        for row_idx in range(5, ws.max_row + 1):
            for col in metrics.keys():
                col_idx = ord(col) - ord('A') + 1
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    if cell.value >= 0.8:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                    elif cell.value >= 0.6:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        
        # Save the updated workbook
        eval_excel_path = excel_path.replace('.xlsx', '_evaluated.xlsx')
        wb.save(eval_excel_path)
        
        logger.info(f"RAGAS evaluation completed and saved to {eval_excel_path}")
        return eval_excel_path
        
    except Exception as e:
        logger.error(f"Error in RAGAS evaluation: {e}")
        logger.error(traceback.format_exc())
        return None
    
